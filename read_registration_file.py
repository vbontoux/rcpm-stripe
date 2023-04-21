
import config
from openpyxl import load_workbook


def read_race_registration_excel_file(filepath):
  #Selecting specific sheet
  wb = load_workbook(filepath)
  sheet = wb.worksheets[3]
  return sheet

sheet = read_race_registration_excel_file(filepath=config.excel_file_path)
#rowlist = list(sheet.rows)
m_row = sheet.max_row
m_col = sheet.max_column

for r in range(2, m_row + 1):
  name = sheet.cell(row = r, column = 1).value
  struct = sheet.cell(row = r, column = 2).value
  resp = sheet.cell(row = r, column = 3).value
  email = sheet.cell(row = r, column = 4).value
  tel = sheet.cell(row = r, column = 5).value
  price = sheet.cell(row = r, column = 10).value.replace(',00 €', '')
  payed = sheet.cell(row = r, column = 12).value
  
  if not payed: 
    print(f'echo "{email}"')
    print(f'python price.py -e {email} -c "{name}" -s "{struct}" -n "{resp}" -t "{tel}" -p "{price}"  | jq -r .id | python link.py | jq -r .url')
